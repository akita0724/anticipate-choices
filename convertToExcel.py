from json import load
from openpyxl import Workbook

from data.data import SECTION

way = 4


fileName = f"./data/way{way}_result.json"


with open(fileName, "r") as f:
    fileData = load(f)


def way2(N: int):
    wb = Workbook()

    ws = wb.active
    ws.title = "result"

    digests = []
    sh = wb.worksheets[0]

    for col, section in enumerate(SECTION):

        digests_tmp = [0] * 10

        sectionData = fileData[section][str(N)]

        sh.cell(1, col * 2 + 2, section)
        for row, pointSt in enumerate(sectionData.keys()):
            point = int(float(pointSt))

            sh.cell(row + 2, col * 2 + 1, point)
            sh.cell(row + 2, col * 2 + 2, sectionData[pointSt])

            digests_tmp[min(point // 10, 9)] += sectionData[pointSt]

        digests.append(digests_tmp)

    wb.create_sheet("digest")
    sh = wb.worksheets[1]

    for col, point in enumerate(digests):
        sh.cell(1, col + 2, SECTION[col])
        for row, section in enumerate(point):
            sh.cell(row + 2, col + 2, section)

    wb.save(f"./data/way2_{N}.xlsx")


def way3_4(N: int, n: int):
    wb = Workbook()

    ws = wb.active
    ws.title = "result"

    digests = []
    sh = wb.worksheets[0]

    for col, section in enumerate(SECTION):

        digests_tmp = [0] * 10

        sectionData = fileData[section][str(N)]

        sh.cell(1, col * 2 + 2, section)
        for row, pointSt in enumerate(sectionData.keys()):
            point = int(float(pointSt))

            sh.cell(row + 2, col * 2 + 1, point)
            sh.cell(row + 2, col * 2 + 2, sectionData[pointSt])

            digests_tmp[min(point // 10, 9)] += sectionData[pointSt]

        digests.append(digests_tmp)

    wb.create_sheet("digest")
    sh = wb.worksheets[1]

    for col, point in enumerate(digests):
        sh.cell(1, col + 2, SECTION[col])
        for row, section in enumerate(point):
            sh.cell(row + 2, col + 2, section)

    wb.save(f"./data/way{n}_{N}.xlsx")


if way == 1:
    wb = Workbook()

    ws = wb.active

    sh = wb.worksheets[0]

    ws.title = (fileName.split("_")[-1]).split(".")[0]

    digests = []

    for col, section in enumerate(SECTION):

        digests_tmp = [0] * 10

        sectionData = fileData[section]
        sh.cell(1, col * 2 + 2, section)
        for row, pointSt in enumerate(sectionData.keys()):
            point = int(float(pointSt))

            sh.cell(row + 2, col * 2 + 1, point)
            sh.cell(row + 2, col * 2 + 2, sectionData[pointSt])

            digests_tmp[min(point // 10, 9)] += sectionData[pointSt]

        digests.append(digests_tmp)

    wb.create_sheet("digest")
    sh = wb.worksheets[1]

    for col, point in enumerate(digests):
        sh.cell(1, col + 2, SECTION[col])
        for row, section in enumerate(point):
            sh.cell(row + 2, col + 2, section)

    wb.save("./data/way1.xlsx")


elif way == 2:
    for N in range(1, 6):
        way2(N)

else:
    for N in range(1, 6):
        way3_4(N, way)
